import openpyxl

global f
f = openpyxl.load_workbook(filename="attendance.xlsx")

def listOfTraineesManip(sheet):
    while True:
        print("OPTIONS:\n1. Add trainee\n2. Delete trainee\n3. Update trainee\n4. View Sheet\n5. Exit")
        choice = int(input("Pick a choice: "))
        if choice == 1:
            flag = True
            id = int(input("Enter the trainee's id: "))
            for cellObj in sheet.rows:
                if id == cellObj[0].value:
                    flag = False
            if flag is True:
                name = input("Enter the trainee's full name: ")
                course = input("Enter the course the trainee is taking: ")
                background = input("Enter the trainee's background (degree/work experience): ")
                lastRow = 0
                for cellObj in sheet.rows:
                    lastRow += 1
                sheet.insert_rows(lastRow+1)
                row = [id, name, course, background]
                for i in range(len(row)):
                    sheet.cell(row=lastRow+1, column=i+1).value = row[i]
                f.save(filename="attendance.xlsx")
            else:
                print("There is currently a trainee with the given ID")

        elif choice == 2:
            toDelete = False
            row = 0
            id = int(input("Enter the existing trainee's id: "))
            for cellObj in sheet.rows:
                row += 1
                if cellObj[0].value == id:
                    toDelete = True

            if toDelete == False:
                print("Trainee with given ID could not be found")
            else:
                sheet.delete_rows(idx=row)
                f.save(filename="attendance.xlsx")

        elif choice == 3:
            toEdit = False
            row = 0
            id = int(input("Enter the existing trainee's id: "))
            for cellObj in sheet.rows:
                row += 1
                if cellObj[0].value == id:
                    toEdit = True

            if toEdit == False:
                print("Trainee with given ID could not be found")
            else:
                name = input("Enter the trainee's name: ")
                course = input("Enter the course the trainee is taking: ")
                background = input("Enter the trainee's background (degree/work experience): ")
                for cellObj in sheet.rows:
                    if cellObj[0].value == id:
                        cellObj[1].value = name
                        cellObj[2].value = course
                        cellObj[3].value = background
                        break
                f.save(filename="attendance.xlsx")

        elif choice == 4:
            print("ID\t\tName\t\t\tCourse\t\tBackground")
            print("--------------------------------------------------------------------------")
            for cellObj in sheet.rows:
                print(f"{cellObj[0].value}\t\t{cellObj[1].value}\t\t\t{cellObj[2].value}\t\t\t{cellObj[3].value}")

        elif choice == 5:
            break
        else:
            print("Incorrect choice")

def listOfTrainersManip(sheet):
    while True:
        print("OPTIONS:\n1. Add trainer\n2. Delete trainer\n3. Update trainer\n4. View Sheet\n5. Exit")
        choice = int(input("Pick a choice: "))
        if choice == 1:
            flag = True
            id = int(input("Enter the trainer's id: "))
            for cellObj in sheet.rows:
                if id == cellObj[0].value:
                    flag = False
            if flag is True:
                name = input("Enter the trainer's full name: ")
                email = input("Enter the trainer's email: ")
                phone = int(input("Enter the trainer's phone number: "))
                course = input("Enter the course that the trainer teaches: ")
                lastRow = 0
                for cellObj in sheet.rows:
                    lastRow += 1
                sheet.insert_rows(lastRow+1)
                row = [id, name, email, phone, course]
                for i in range(len(row)):
                    sheet.cell(row=lastRow+1, column=i+1).value = row[i]
                f.save(filename="attendance.xlsx")
            else:
                print("There is currently a trainer with the given ID")

        elif choice == 2:
            toDelete = False
            row = 0
            id = int(input("Enter the existing trainer's id: "))
            for cellObj in sheet.rows:
                row += 1
                if cellObj[0].value == id:
                    toDelete = True

            if toDelete == False:
                print("Trainer with given ID could not be found")
            else:
                sheet.delete_rows(idx=row)
                f.save(filename="attendance.xlsx")

        elif choice == 3:
            toEdit = False
            row = 0
            id = int(input("Enter the existing trainer's id: "))
            for cellObj in sheet.rows:
                row += 1
                if cellObj[0].value == id:
                    toEdit = True

            if toEdit == False:
                print("Trainer with given ID could not be found")
            else:
                name = input("Enter the trainer's full name: ")
                email = input("Enter the trainer's email: ")
                phone = int(input("Enter the trainer's phone number: "))
                course = input("Enter the course that the trainer teaches: ")
                for cellObj in sheet.rows:
                    if cellObj[0].value == id:
                        cellObj[1].value = name
                        cellObj[2].value = email
                        cellObj[3].value = phone
                        cellObj[4].value = course
                        break
                f.save(filename="attendance.xlsx")

        elif choice == 4:
            print("ID\t\tName\t\t\tEmail\t\t\tPhone number\t\tCourse")
            print("---------------------------------------------------------------------------------------")
            for cellObj in sheet.rows:
                print(f"{cellObj[0].value}\t\t{cellObj[1].value}\t\t\t{cellObj[2].value}\t\t\t{cellObj[3].value}\t\t\t{cellObj[4].value}")

        elif choice == 5:
            break
        else:
            print("Incorrect choice")

def listOfManagersManip(manager, trainer):
    while True:
        print("OPTIONS:\n1. Add trainer\n2. Delete trainer\n3. Update trainer\n4. View Sheet\n5. Exit")
        choice = int(input("Pick a choice: "))
        if choice == 1:
            flag = False
            id = int(input("Enter the manager's id: "))
            name = input("Enter the manager's full name: ")
            email = input("Enter the manager's email: ")
            phone = int(input("Enter the manager's phone number: "))
            trainerid = input("Enter the trainer ID that the manager is overseeing (N/A if not managing anyone): ")
            for cellObj in trainer.rows:
                if trainerid == cellObj[0].value:
                    flag = True
            if flag is True or trainerid == "N/A":
                lastRow = 0
                for cellObj in manager.rows:
                    lastRow += 1
                manager.insert_rows(lastRow+1)
                row = [id, name, email, phone]
                for i in range(len(row)):
                    manager.cell(row=lastRow+1, column=i+1).value = row[i]
                f.save(filename="attendance.xlsx")
            else:
                print("Manager ID could not be found")

        elif choice == 2:
            toDelete = False
            row = 0
            id = int(input("Enter the existing manager's id: "))
            for cellObj in manager.rows:
                row += 1
                if cellObj[0].value == id:
                    toDelete = True

            if toDelete == False:
                print("Manager with given ID could not be found")
            else:
                manager.delete_rows(idx=row)
                f.save(filename="attendance.xlsx")

        elif choice == 3:
            flag = False
            toEdit = False
            row = 0
            id = int(input("Enter the existing manager's id: "))
            for cellObj in manager.rows:
                row += 1
                if cellObj[0].value == id:
                    toEdit = True

            if toEdit == False:
                print("Manager with given ID could not be found")
            else:
                name = input("Enter the manager's full name: ")
                email = input("Enter the manager's email: ")
                phone = int(input("Enter the manager's phone number: "))
                trainerid = input("Enter the trainer ID that the manager is overseeing (N/A if not managing anyone): ")
                for cellObj in trainer.rows:
                    if trainerid == cellObj[0].value:
                        flag = True
                for cellObj in manager.rows:
                    if cellObj[0].value == id:
                        cellObj[1].value = name
                        cellObj[2].value = email
                        cellObj[3].value = phone
                        break
                f.save(filename="attendance.xlsx")

        elif choice == 4:
            print("ID\t\tName\t\t\tEmail\t\t\tPhone number")
            print("---------------------------------------------------------------------------------------")
            for cellObj in manager.rows:
                print(f"{cellObj[0].value}\t\t{cellObj[1].value}\t\t\t{cellObj[2].value}\t\t\t{cellObj[3].value}")

        elif choice == 5:
            break
        else:
            print("Incorrect choice")

def courseDetailsManip(sheet):
    while True:
        print("OPTIONS:\n1. Add Course\n2. Delete Course\n3. Update course\n4. View Sheet\n5. Exit")
        choice = int(input("Pick a choice: "))
        if choice == 1:
            id = input("Enter the course's id: ")
            description = input("Enter what the course is: ")
            lastRow = 0
            for cellObj in sheet.rows:
                lastRow += 1
            sheet.insert_rows(lastRow+1)
            row = [id, description]
            for i in range(len(row)):
                sheet.cell(row=lastRow+1, column=i+1).value = row[i]
            f.save(filename="attendance.xlsx")

        elif choice == 2:
            toDelete = False
            row = 0
            id = input("Enter the existing course's id: ")
            for cellObj in sheet.rows:
                row += 1
                if cellObj[0].value == id:
                    toDelete = True

            if toDelete == False:
                print("Course with given ID could not be found")
            else:
                sheet.delete_rows(idx=row)
                f.save(filename="attendance.xlsx")

        elif choice == 3:
            toEdit = False
            row = 0
            id = int(input("Enter the existing trainee's id: "))
            for cellObj in sheet.rows:
                row += 1
                if cellObj[0].value == id:
                    toEdit = True

            if toEdit == False:
                print("Course with given ID could not be found")
            else:
                description = input("Enter what the course is: ")
                for cellObj in sheet.rows:
                    if cellObj[0].value == id:
                        cellObj[1].value = description
                        break
                f.save(filename="attendance.xlsx")

        elif choice == 4:
            print("ID\tDescription")
            print("--------------------------------------------------------------------------")
            for cellObj in sheet.rows:
                print(f"{cellObj[0].value}\t{cellObj[1].value}")

        elif choice == 5:
            break
        else:
            print("Incorrect choice")

def mapCourseTrainee(map, course, trainee):
    while True:
        print("OPTIONS:\n1. Map course and trainees\n2. View Sheet\n3. Exit")
        choice = int(input("Pick a choice: "))
        if choice == 1:
            row = 0
            for courseObj in course.rows:
                row += 1
                rowlist = []
                for traineeObj in trainee.rows:
                    if courseObj[0].value == traineeObj[2].value:
                        rowlist.append(traineeObj[0].value)
                map.insert_rows(row)
                map.cell(row=row, column=1).value = courseObj[0].value
                for i in range(len(rowlist)):
                    if map.cell(row=row, column=2).value is None:
                        map.cell(row=row, column=2).value = str(rowlist[i])
                    else:
                        map.cell(row=row, column=2).value += ", " + str(rowlist[i])
            f.save(filename="attendance.xlsx")
        elif choice == 2:
            print("Course ID\tTrainee IDs")
            print("--------------------------------------------------------------------------")
            for mapObj in map.rows:
                print(f"{mapObj[0].value}\t\t{mapObj[1].value}")
        elif choice == 3:
            break

def mapCourseTrainer(map, course, trainer):
    while True:
        print("OPTIONS:\n1. Map course and trainees\n2. View Sheet\n3. Exit")
        choice = int(input("Pick a choice: "))
        if choice == 1:
            row = 0
            for courseObj in course.rows:
                row += 1
                rowlist = []
                for trainerObj in trainer.rows:
                    if courseObj[0].value == trainerObj[4].value:
                        rowlist.append(trainerObj[0].value)
                map.insert_rows(row)
                map.cell(row=row, column=1).value = courseObj[0].value
                for i in range(len(rowlist)):
                    if map.cell(row=row, column=2).value is None:
                        map.cell(row=row, column=2).value = str(rowlist[i])
                    else:
                        map.cell(row=row, column=2).value += ", " + str(rowlist[i])
            f.save(filename="attendance.xlsx")
        elif choice == 2:
            print("Course ID\tTrainee IDs")
            print("--------------------------------------------------------------------------")
            for mapObj in map.rows:
                print(f"{mapObj[0].value}\t\t{mapObj[1].value}")
        elif choice == 3:
            break

def createSheet(trainees, trainers):
    while True:
        starttime = input("Input the start time of the course: ")
        endtime = input("Input the end time of the course: ")
        course = input("Enter the course: ")
        trainer = None
        temp = int(input("Enter the ID of the trainer who will be teaching this session: "))
        for cellObj in trainers.rows:
            if temp == cellObj[0].value:
                if cellObj[4].value == course:
                    trainer = cellObj
        if trainer is not None:
            print("New Session")
            print(f"Start Time: {starttime}    End Time: {endtime}")
            print(f"Trainer: {trainer[1].value}")
            choice = input("Are these details correct? Enter Y or N: ")
            if choice == "Y":
                print("Trainees in cohort")
                print("----------------------------")
                for cellObj in trainees.rows:
                    if cellObj[2].value == course:
                        print(f"{cellObj[0].value}\t{cellObj[1].value}\tP")
                absentees = input("Enter the trainees that are absent, separate with a comma: ")
                absentees = absentees.split(",")
                print("Sheet that will be saved")
                print("Trainees in cohort")
                print("----------------------------")
                for cellObj in trainees.rows:
                    if cellObj[2].value == course:
                        if str(cellObj[0].value) in absentees:
                            print(f"{cellObj[0].value}\t{cellObj[1].value}\tA")
                        else:
                            print(f"{cellObj[0].value}\t{cellObj[1].value}\tP")
                choice = input("Do you wish to save this? Enter Y or N: ")
                if choice == "Y":
                    date = input("Enter the date (format MMDD_YYYY): ")
                    newsheet = f.create_sheet(date)
                    newsheet.insert_rows(3)
                    newsheet.cell(row=1, column=1).value = starttime
                    newsheet.cell(row=1, column=2).value = endtime
                    newsheet.cell(row=2, column=1).value = "Trainer"
                    newsheet.cell(row=2, column=2).value = trainer[1].value
                    newsheet.cell(row=3, column=1).value = "Attendance of trainees"
                    lastRow = 0
                    for cellObj in newsheet.rows:
                        lastRow += 1
                    newsheet.insert_rows(lastRow + 1)
                    for cellObj in trainees.rows:
                        newsheet.insert_rows(lastRow + 1)
                        if cellObj[2].value == course:
                            if str(cellObj[0].value) in absentees:
                                row = [cellObj[0].value, cellObj[1].value, "A"]
                                for i in range(len(row)):
                                    newsheet.cell(row=lastRow + 1, column=i + 1).value = row[i]
                            else:
                                row = [cellObj[0].value, cellObj[1].value, "P"]
                                for i in range(len(row)):
                                    newsheet.cell(row=lastRow + 1, column=i + 1).value = row[i]
                        lastRow += 1
                    f.save(filename="attendance.xlsx")
                    break
            else:
                print("Restarting...")





def openSheet(sheetname):
    return f[sheetname]

def menu():
    i = 1
    sheet = None
    while True:
        for name in f.sheetnames:
            print(f"{i}. {str(f[name])}")
            i += 1
        print(f"{i}. Create attendance sheet")
        print(f"8. Exit")
        choice = int(input("Choose what you wish to do: "))
        if choice == 1:
            sheet = openSheet("ListOfTrainees")
            listOfTraineesManip(sheet)
        elif choice == 2:
            sheet = openSheet("ListOfTrainers")
            listOfTrainersManip(sheet)
        elif choice == 3:
            sheet1 = openSheet("ListOfManagers")
            sheet2 = openSheet("ListOfTrainers")
            listOfManagersManip(sheet1, sheet2)
        elif choice == 4:
            sheet = openSheet("CourseDetails")
            courseDetailsManip(sheet)
        elif choice == 5:
            sheet1 = openSheet("MappingCourseTrainees")
            sheet2 = openSheet("CourseDetails")
            sheet3 = openSheet("ListOfTrainees")
            mapCourseTrainee(sheet1, sheet2, sheet3)
        elif choice == 6:
            sheet1 = openSheet("MappingCourseTrainers")
            sheet2 = openSheet("CourseDetails")
            sheet3 = openSheet("ListOfTrainers")
            mapCourseTrainer(sheet1, sheet2, sheet3)
        elif choice == 7:
            sheet1 = openSheet("ListOfTrainees")
            sheet2 = openSheet("ListOfTrainers")
            createSheet(sheet1, sheet2)
        elif choice == 8:
            break
        else:
            print("Incorrect choice")
        i = 1

menu()
